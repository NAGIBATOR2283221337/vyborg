from datetime import time, datetime
from backend.processors.time_transfer import coerce_time, write_time_cell, process_time_columns
from openpyxl import Workbook
import tempfile, os


def test_coerce_time_basic():
    assert coerce_time(time(6,0,15)).strftime('%H:%M:%S') == '06:00:15'
    assert coerce_time(datetime(2025,1,1,7,8,9)).strftime('%H:%M:%S') == '07:08:09'
    assert coerce_time('06:05').strftime('%H:%M:%S') == '06:05:00'
    assert coerce_time('06:05:10').strftime('%H:%M:%S') == '06:05:10'
    # excel serial 0.5 суток -> 12:00:00
    assert coerce_time(0.5).strftime('%H:%M:%S') == '12:00:00'


def test_write_time_cell():
    wb = Workbook(); ws = wb.active
    c = ws.cell(1,1)
    write_time_cell(c, time(1,2,3))
    assert c.value.strftime('%H:%M:%S') == '01:02:03'
    assert c.number_format.lower() == 'hh:mm:ss'


def test_process_time_columns(tmp_path):
    # создаём временную книгу с заголовком на 10-й строке
    wb = Workbook(); ws = wb.active; ws.title = 'Лист1'
    for r in range(1, 15):
        ws.cell(r,1).value = ''
    ws.cell(10,3).value = 'Время (часы, мин.)'
    ws.cell(11,3).value = '06:00:15'
    ws.cell(12,3).value = 0.5  # 12:00:00
    ws.cell(13,3).value = '07:30'
    ws.cell(14,3).value = None
    book_path = tmp_path/'sample.xlsx'
    wb.save(book_path)
    stats = process_time_columns(str(book_path), sheet='Лист1', header_row=10, default_time=time(0,0,0))
    assert stats['before_time_empty'] == 1
    assert stats['after_time_empty'] == 0
    assert stats['filled_cells'] == 1
    assert os.path.exists(stats['saved_to'])

