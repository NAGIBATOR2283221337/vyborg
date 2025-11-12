# debug_time_probe.py
from openpyxl import load_workbook
import re
from datetime import datetime, date, time
from openpyxl.utils.datetime import from_excel as excel_from_serial

HDR_ROW = 10
SHEET = "Лист1"

def find_col(ws, header_row, pattern):
    pat = re.compile(pattern, re.I)
    for c in range(1, ws.max_column+1):
        name = str(ws.cell(header_row, c).value or "")
        if pat.search(name): return c, name
    return None, None

def coerce_time(value, number_format=None):
    # Поддержка time, datetime, excel-числа, строки HH:MM[:SS]
    if value is None or value == "": return None
    if isinstance(value, time): return value
    if isinstance(value, datetime): return value.time()
    if isinstance(value, (int, float)):
        try:
            dt = excel_from_serial(value)  # 1899-12-30 origin
            return dt.time()
        except Exception:
            pass
    s = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None

# 1) Проверим, что приёмник действительно имеет колонку ВРЕМЕНИ
dst = "Копия Сентябрь в работе.xlsx"           # <-- подставь свой файл
wb = load_workbook(dst, data_only=True)
ws = wb[SHEET]
c_time, hdr_time = find_col(ws, HDR_ROW, r"^время(?:\b|\s*\()")
c_date, hdr_date = find_col(ws, HDR_ROW, r"^дата(\b|\s*\()")
c_title, hdr_title = find_col(ws, HDR_ROW, r"^наименование")

print("TIME COL:", c_time, hdr_time)
print("DATE COL:", c_date, hdr_date)
print("TITLE COL:", c_title, hdr_title)

# 2) Показать 5 ячеек «Время» ДО записи
for r in range(HDR_ROW+1, HDR_ROW+6):
    v = ws.cell(r, c_time).value if c_time else None
    print(f"Row {r}: raw time cell =", repr(v))

# 3) Проба чтения времени из «источника»
src = "Копия Сентябрь в работе.xlsx"           # <-- подставь свой файл
wb2 = load_workbook(src, data_only=True)
ws2 = wb2[SHEET]
# Попробуем найти колонку с временем/датой
c_dt, _ = find_col(ws2, HDR_ROW, r"дата\s*и\s*время")
c_t, _  = find_col(ws2, HDR_ROW, r"^время(?:\b|\s*\()")
print("SRC DT:", c_dt, "SRC TIME:", c_t)

for r in range(HDR_ROW+1, HDR_ROW+6):
    raw = ws2.cell(r, (c_t or c_dt)).value if (c_t or c_dt) else None
    coerced = coerce_time(raw)
    print(f"SRC Row {r}: raw={repr(raw)} -> coerced={coerced}")

print("Если TIME COL = None, значит паттерн заголовка не подошёл. Если coerced=None при нормальных значениях — шлёпни сюда вывод.")


