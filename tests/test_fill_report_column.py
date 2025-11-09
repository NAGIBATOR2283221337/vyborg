import io
from datetime import datetime
from openpyxl import Workbook, load_workbook

from backend.processors.write_report_openpyxl import fill_report_column, match_times_for_title


def make_schedule_bytes():
    wb = Workbook(); ws = wb.active
    ws.title = 'Sheet1'
    ws.cell(1,2).value = 'Понедельник, 1 сентября 2025'
    ws.cell(2,1).value = '06:00'; ws.cell(2,2).value = 'Новости'
    ws.cell(3,1).value = '08:00'; ws.cell(3,2).value = 'Гора самоцветов 63,64'
    bio = io.BytesIO(); wb.save(bio); bio.seek(0); return bio.getvalue()


def make_report_file(path: str):
    wb = Workbook(); ws = wb.active
    ws.title = 'росийские произведения'
    # Header row
    ws.cell(1,1).value = 'Наименование аудиовизуального произведения (номер и название серии)'
    ws.cell(1,2).value = 'Дата и время выхода в эфир (число, часы, мин.)'
    # Data rows
    ws.cell(2,1).value = 'Гора самоцветов. 63 серия'
    ws.cell(3,1).value = 'Новости'
    ws.cell(4,1).value = 'Несуществующая'
    wb.save(path)


def test_fill_report_column(tmp_path):
    schedule_bytes = make_schedule_bytes()
    report_path = tmp_path / 'report.xlsx'
    make_report_file(str(report_path))

    fill_report_column(schedule_bytes, str(report_path))

    wb = load_workbook(str(report_path)); ws = wb['росийские произведения']
    val1 = ws.cell(2,2).value
    val2 = ws.cell(3,2).value
    # Проверяем формат
    assert '01.09.2025' in val1 and '08:00' in val1
    assert '06:00' in val2
    # Несуществующая не тронута
    assert ws.cell(4,2).value is None
    wb.close()

if __name__ == '__main__':
    import tempfile, pathlib
    tmp = tempfile.TemporaryDirectory()
    p = pathlib.Path(tmp.name)
    test_fill_report_column(p)
    print('✅ fill_report_column интеграционный тест пройден')

