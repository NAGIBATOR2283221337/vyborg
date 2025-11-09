"""
Создание примеров файлов для тестирования системы
"""
import pandas as pd
import openpyxl
from openpyxl import Workbook
import os

def create_sample_schedule():
    """Создает пример файла сетки"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Сетка"

    # Добавляем данные сетки
    data = [
        ["Понедельник, 11 ноября 2025", ""],
        ["10:00", "Утренние новости"],
        ["10:30", "Доброе утро, страна!"],
        ["12:00", "Дневные новости"],
        ["14:00", "Ток-шоу 'Жизнь замечательных людей'"],
        ["16:00", "Документальный фильм 'Природа России'"],
        ["18:00", "Вечерние новости"],
        ["19:00", "Сериал 'Московские тайны' 15 серия"],
        ["21:00", "Главные новости дня"],
        ["", ""],
        ["Вторник, 12 ноября 2025", ""],
        ["10:00", "Утренние новости"],
        ["10:30", "Доброе утро, страна!"],
        ["12:00", "Дневные новости"],
        ["14:00", "Ток-шоу 'Время говорить'"],
        ["16:00", "Художественный фильм 'Офицеры'"],
        ["18:00", "Вечерние новости"],
        ["19:00", "Сериал 'Московские тайны' 16 серия"],
        ["21:00", "Главные новости дня"],
    ]

    for row_idx, (col_a, col_b) in enumerate(data, 1):
        ws.cell(row_idx, 1, col_a)
        ws.cell(row_idx, 2, col_b)

    return wb

def create_sample_report_rus():
    """Создает пример российского отчёта"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    # Заголовки
    headers = ["№", "Наименование аудиовизуального произведения", "Дата и время показов"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(1, col_idx, header)

    # Данные отчёта
    data = [
        [1, "Утренние новости", "11.11.2025"],
        [2, "Доброе утро страна", "11.11.2025"],
        [3, "Дневные новости программа", "11.11.2025"],
        [4, "Жизнь замечательных людей ток-шоу", "11.11.2025"],
        [5, "Природа России документальный", "11.11.2025"],
        [6, "Вечерние новости", "11.11.2025"],
        [7, "Московские тайны сериал", "11.11.2025"],
        [8, "Главные новости", "11.11.2025"],
        [9, "Время говорить", "12.11.2025"],
        [10, "Офицеры фильм", "12.11.2025"],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, value)

    return wb

def create_sample_report_foreign():
    """Создает пример иностранного отчёта"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    # Заголовки
    headers = ["№", "Название передачи", "Дата показа"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(1, col_idx, header)

    # Данные отчёта
    data = [
        [1, "Morning News", "11.11.2025"],
        [2, "Good Morning Country", "11.11.2025"],
        [3, "Daily News", "11.11.2025"],
        [4, "Talk Show Life", "11.11.2025"],
        [5, "Nature Documentary", "11.11.2025"],
        [6, "Evening News", "11.11.2025"],
        [7, "Moscow Mysteries Series", "11.11.2025"],
        [8, "Main News", "11.11.2025"],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, value)

    return wb

if __name__ == "__main__":
    # Создаем папку для примеров
    examples_dir = "examples"
    if not os.path.exists(examples_dir):
        os.makedirs(examples_dir)

    # Создаем примеры файлов
    schedule_wb = create_sample_schedule()
    schedule_wb.save(os.path.join(examples_dir, "sample_schedule.xlsx"))
    print("✓ Создан пример сетки: examples/sample_schedule.xlsx")

    rus_report_wb = create_sample_report_rus()
    rus_report_wb.save(os.path.join(examples_dir, "sample_report_rus.xlsx"))
    print("✓ Создан пример российского отчёта: examples/sample_report_rus.xlsx")

    foreign_report_wb = create_sample_report_foreign()
    foreign_report_wb.save(os.path.join(examples_dir, "sample_report_foreign.xlsx"))
    print("✓ Создан пример иностранного отчёта: examples/sample_report_foreign.xlsx")

    print("\nПримеры файлов созданы! Теперь можно тестировать систему.")
    print("Файлы находятся в папке 'examples/'")
