#!/usr/bin/env python3
"""
Тест исправления PermissionError для системы обработки отчетов
"""
import os
import sys
import tempfile
import time
from pathlib import Path

# Добавляем пути
current_dir = Path(__file__).parent
sys.path.insert(0, str(current_dir))
sys.path.insert(0, str(current_dir / "backend"))

def create_test_files():
    """Создает простые тестовые Excel файлы"""
    from openpyxl import Workbook

    # Создаем сетку
    schedule_wb = Workbook()
    ws = schedule_wb.active

    # Простые тестовые данные
    ws.cell(1, 1, "Понедельник, 11 ноября 2025")
    ws.cell(2, 1, "10:00")
    ws.cell(2, 2, "Тестовая программа")
    ws.cell(3, 1, "12:00")
    ws.cell(3, 2, "Другая программа")

    # Создаем отчет
    report_wb = Workbook()
    ws = report_wb.active

    ws.cell(1, 1, "№")
    ws.cell(1, 2, "Наименование аудиовизуального произведения")
    ws.cell(1, 3, "Дата и время показов")

    ws.cell(2, 1, 1)
    ws.cell(2, 2, "Тестовая программа")
    ws.cell(2, 3, "11.11.2025")

    return schedule_wb, report_wb

def test_permission_error_fix():
    """Тестирует исправление PermissionError"""
    print("🧪 Тест исправления PermissionError")
    print("=" * 50)

    try:
        # Импортируем модули
        print("📥 Импорт модулей...")
        from backend.processors import processor_rus
        print("✅ Модули импортированы")

        # Создаем тестовые файлы
        print("📄 Создание тестовых файлов...")
        schedule_wb, report_wb = create_test_files()

        # Конвертируем в байты
        from io import BytesIO

        schedule_buffer = BytesIO()
        schedule_wb.save(schedule_buffer)
        schedule_bytes = schedule_buffer.getvalue()
        schedule_buffer.close()
        schedule_wb.close()

        report_buffer = BytesIO()
        report_wb.save(report_buffer)
        report_bytes = report_buffer.getvalue()
        report_buffer.close()
        report_wb.close()

        print("✅ Тестовые файлы созданы")

        # Параметры обработки
        params = {
            'max_shows': 3,
            'fuzzy_cutoff': 0.20,
            'min_token_overlap': 0.35,
            'delete_unmatched': True
        }

        print("⚙️  Запуск обработки...")

        # Пробуем обработать несколько раз подряд
        for i in range(3):
            print(f"   Попытка {i+1}/3...")

            try:
                result_bytes = processor_rus.process(schedule_bytes, report_bytes, params)
                print(f"   ✅ Попытка {i+1} успешна, размер результата: {len(result_bytes)} байт")

                # Небольшая пауза между попытками
                time.sleep(0.5)

            except PermissionError as e:
                print(f"   ❌ Попытка {i+1}: PermissionError - {e}")
                return False
            except Exception as e:
                print(f"   ❌ Попытка {i+1}: Другая ошибка - {e}")
                return False

        print("🎉 ВСЕ ПОПЫТКИ УСПЕШНЫ!")
        print("✅ PermissionError исправлена!")
        return True

    except Exception as e:
        print(f"❌ Ошибка теста: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    if test_permission_error_fix():
        print("\n🎉 ТЕСТ ПРОЙДЕН!")
        print("Система готова к работе без ошибок PermissionError")
    else:
        print("\n❌ ТЕСТ НЕ ПРОЙДЕН!")
        print("Проблема с PermissionError еще не решена")

    print("=" * 50)
