#!/usr/bin/env python3
"""
Интеграционный тест системы обработки отчётов
Создает тестовые файлы и проверяет полный цикл обработки
"""
import os
import sys
import tempfile
from pathlib import Path

# Добавляем путь к backend
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

def create_test_files():
    """Создает тестовые файлы в временной директории"""
    temp_dir = tempfile.mkdtemp(prefix="report_test_")

    # Создаем сетку
    from openpyxl import Workbook

    # Сетка
    schedule_wb = Workbook()
    ws = schedule_wb.active

    schedule_data = [
        ["Понедельник, 11 ноября 2025", ""],
        ["10:00", "Утренние новости"],
        ["12:00", "Дневник программы"],
        ["14:00", "Документальный фильм о природе"],
        ["16:00", "Ток-шоу Время разговора"],
        ["18:00", "Вечерние новости"],
        ["20:00", "Художественный фильм Офицеры"],
    ]

    for row_idx, (col_a, col_b) in enumerate(schedule_data, 1):
        ws.cell(row_idx, 1, col_a)
        ws.cell(row_idx, 2, col_b)

    schedule_path = os.path.join(temp_dir, "test_schedule.xlsx")
    schedule_wb.save(schedule_path)

    # Отчёт
    report_wb = Workbook()
    ws = report_wb.active

    # Заголовки
    ws.cell(1, 1, "№")
    ws.cell(1, 2, "Наименование аудиовизуального произведения")
    ws.cell(1, 3, "Дата и время показов")

    # Данные
    report_data = [
        [1, "Утренние новости программа", "11.11.2025"],
        [2, "Дневник программы телевидения", "11.11.2025"],
        [3, "Документальный фильм природа", "11.11.2025"],
        [4, "Ток-шоу время для разговора", "11.11.2025"],
        [5, "Вечерние новости", "11.11.2025"],
        [6, "Художественный фильм офицеры", "11.11.2025"],
        [7, "Неизвестная программа", "11.11.2025"],  # Эта должна быть удалена
    ]

    for row_idx, row_data in enumerate(report_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row_idx, col_idx, value)

    report_path = os.path.join(temp_dir, "test_report.xlsx")
    report_wb.save(report_path)

    return temp_dir, schedule_path, report_path

def test_processing():
    """Тестирует обработку файлов"""
    print("🧪 Создание тестовых файлов...")
    temp_dir, schedule_path, report_path = create_test_files()

    try:
        print("📂 Тестовые файлы созданы в:", temp_dir)

        # Импортируем процессор
        from processors import processor_rus

        # Читаем файлы
        with open(schedule_path, 'rb') as f:
            schedule_bytes = f.read()

        with open(report_path, 'rb') as f:
            report_bytes = f.read()

        # Параметры обработки
        params = {
            'max_shows': 3,
            'fuzzy_cutoff': 0.20,
            'min_token_overlap': 0.35,
            'delete_unmatched': True
        }

        print("⚙️  Запуск обработки...")

        # Обрабатываем
        result_bytes = processor_rus.process(schedule_bytes, report_bytes, params)

        # Сохраняем результат
        result_path = os.path.join(temp_dir, "result.xlsx")
        with open(result_path, 'wb') as f:
            f.write(result_bytes)

        print(f"✅ Обработка завершена! Результат сохранен: {result_path}")

        # Проверяем результат
        from openpyxl import load_workbook
        wb = load_workbook(result_path)
        ws = wb.active

        print("\n📊 Результат обработки:")
        print("Строка | Название | Показы")
        print("-" * 60)

        matched_count = 0
        for row in range(2, ws.max_row + 1):
            title = ws.cell(row, 2).value
            shows = ws.cell(row, 3).value

            if title and shows:
                shows_text = str(shows)[:50] + "..." if len(str(shows)) > 50 else str(shows)
                print(f"{row-1:4d} | {title[:25]:<25} | {shows_text}")
                if "11.11.2025 в" in str(shows):
                    matched_count += 1

        print(f"\n📈 Статистика:")
        print(f"   Всего строк обработано: {ws.max_row - 1}")
        print(f"   Строк с совпадениями: {matched_count}")
        print(f"   Строк удалено: {7 - (ws.max_row - 1)}")  # 7 было в исходном отчёте

        wb.close()

        if matched_count > 0:
            print("\n🎉 Тест пройден успешно!")
            return True
        else:
            print("\n❌ Тест не пройден: нет найденных совпадений")
            return False

    except Exception as e:
        print(f"\n❌ Ошибка во время теста: {e}")
        import traceback
        traceback.print_exc()
        return False

    finally:
        # Очистка
        import shutil
        try:
            shutil.rmtree(temp_dir)
            print(f"🧹 Временные файлы удалены: {temp_dir}")
        except:
            print(f"⚠️  Не удалось удалить временные файлы: {temp_dir}")

def test_web_interface():
    """Тестирует доступность веб-интерфейса"""
    try:
        # Импортируем main модуль
        from main import app
        print("✅ FastAPI приложение загружено успешно")

        # Проверяем, что статические файлы на месте
        frontend_path = Path("frontend")
        required_files = ["index.html", "app.js", "styles.css"]

        for file_name in required_files:
            file_path = frontend_path / file_name
            if file_path.exists():
                print(f"✅ {file_name} найден")
            else:
                print(f"❌ {file_name} не найден")
                return False

        print("🌐 Веб-интерфейс готов к работе")
        return True

    except Exception as e:
        print(f"❌ Ошибка веб-интерфейса: {e}")
        return False

if __name__ == "__main__":
    print("=" * 60)
    print("🚀 ИНТЕГРАЦИОННЫЙ ТЕСТ СИСТЕМЫ ОБРАБОТКИ ОТЧЁТОВ")
    print("=" * 60)

    tests = [
        ("Обработка файлов", test_processing),
        ("Веб-интерфейс", test_web_interface),
    ]

    passed = 0
    for test_name, test_func in tests:
        print(f"\n🔍 Тест: {test_name}")
        print("-" * 40)

        if test_func():
            passed += 1
            print(f"✅ {test_name}: ПРОЙДЕН")
        else:
            print(f"❌ {test_name}: НЕ ПРОЙДЕН")

    print("\n" + "=" * 60)
    print(f"📊 РЕЗУЛЬТАТ: {passed}/{len(tests)} тестов пройдено")

    if passed == len(tests):
        print("🎉 ВСЕ ТЕСТЫ ПРОЙДЕНЫ! Система готова к работе.")
        print("\n🚀 Для запуска выполните:")
        print("   python -m uvicorn backend.main:app --reload --port 8000")
        print("   или запустите start_server.bat")
    else:
        print("⚠️  Некоторые тесты не пройдены. Проверьте ошибки выше.")

    print("=" * 60)
