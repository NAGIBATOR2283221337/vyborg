# strict_match.py – строгий индекс и перенос по (base, episode, date)

from openpyxl import load_workbook
from datetime import datetime, date, timedelta, time
import io
import re
from typing import Dict, Tuple, Optional, Union, List

# Дополнительные регулярки для улучшенного парсинга эпизода
_EP_ANY_RE = re.compile(r'(\d{1,3})\s*(?:серия|выпуск|эпизод|часть)\b', re.I)
_LEADING_CODE_RE = re.compile(r'^\d{4,}[ _-]+')
_EP_SINGLE_RE = re.compile(r'(?:^|[\s.])(\d{1,3})\s*(?:серия|выпуск|эпизод|часть)?$', re.I)

# Нормализация строки: удаление лишних пробелов, перевод в нижний регистр
def norm(s: str) -> str:
    return ' '.join(s.strip().lower().split())

# Парсинг времени в формате ЧЧ:ММ или ЧЧ:ММ:СС
def _parse_time(time_val: Union[str, float]) -> Optional[time]:
    if isinstance(time_val, float):
        # Excel хранит долю суток как число; преобразуем в секунды
        if -1.0 < time_val < 3.0:
            total_sec = int(round(time_val * 24 * 3600)) % (24*3600)
            h = total_sec // 3600; m = (total_sec % 3600)//60; s = total_sec % 60
            return time(h, m, s)
        return None
    if isinstance(time_val, str):
        s = norm(time_val)
        for fmt in ('%H:%M:%S', '%H:%M'):
            try:
                tm = datetime.strptime(s, fmt).time()
                return tm
            except Exception:
                continue
    return None

# Парсинг продолжительности в формате ЧЧ:ММ или ММ:СС
def _parse_duration(duration_str: Union[str, float]) -> Optional[timedelta]:
    if isinstance(duration_str, float):
        # Если продолжительность в формате числа, конвертируем в timedelta
        return timedelta(minutes=duration_str)
    if isinstance(duration_str, str):
        duration_str = norm(duration_str)
        for fmt in ('%H:%M', '%M:%S'):
            try:
                t = datetime.strptime(duration_str, fmt)
                return timedelta(hours=t.hour, minutes=t.minute, seconds=t.second)
            except Exception:
                continue
    return None

# Парсинг даты из заголовка отчета
def _parse_report_date(date_val) -> Optional[date]:
    if isinstance(date_val, str):
        date_val = norm(date_val)
        for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(date_val, fmt).date()
            except Exception:
                continue
    return None

# Парсинг даты из заголовка таблицы
def _parse_header_date(title_cell) -> Optional[date]:
    if isinstance(title_cell, str):
        title_cell = norm(title_cell)
        for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(title_cell, fmt).date()
            except Exception:
                continue
    return None

# Разделение базового названия и множества эпизодов
def split_base_episodes(title: str) -> Tuple[str, List[Optional[int]]]:
    title = norm(title)
    # Поиск всех эпизодов в названии
    eps = _EP_ANY_RE.findall(title)
    eps_set = set()
    for ep in eps:
        try:
            eps_set.add(int(ep))
        except Exception:
            continue
    # Удаление эпизодов из названия для получения базового имени
    base = _EP_ANY_RE.sub('', title).strip()
    return base, list(eps_set)

KeyType = Tuple[str, Optional[int], date]
ValueType = Tuple[datetime, Optional[timedelta]]
# Теперь индекс будет хранить список показов
IndexType = Dict[KeyType, List[ValueType]]

# Разделение названия и эпизода
def split_title_episode(title: str) -> Tuple[str, Optional[int]]:
    # удаляем ведущий числовой код/id если есть
    cleaned = _LEADING_CODE_RE.sub('', title)
    base, eps_set = split_base_episodes(cleaned)
    # Если ровно один эпизод через стандартный парсер
    if len(eps_set) == 1:
        return base, next(iter(eps_set))
    # Попытка найти одиночный эпизод по явному шаблону число+слово
    matches = _EP_ANY_RE.findall(norm(cleaned))
    if len(matches) == 1:
        try:
            return base, int(matches[0])
        except Exception:
            pass
    # Попытка fallback: одиночное число в конце
    m = _EP_SINGLE_RE.search(norm(cleaned))
    if m and not eps_set:
        try:
            return base, int(m.group(1))
        except Exception:
            return base, None
    return base, None

# Построение индекса расписания
def build_schedule_index(xls_bytes: bytes) -> Tuple[IndexType, int]:
    # Возвращает индекс {(base, episode, date): [ (air_dt, duration), ... ]} и число коллизий.
    bio = io.BytesIO(xls_bytes)
    wb = load_workbook(bio, data_only=True)
    index: IndexType = {}
    collisions = 0
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            cur_date: Optional[date] = None
            for r in range(1, ws.max_row + 1):
                title_cell = ws.cell(r, 2).value
                d = _parse_header_date(title_cell)
                if d:
                    cur_date = d
                    continue
                if cur_date is None:
                    continue
                time_cell = ws.cell(r, 1).value
                dur_cell = ws.cell(r, 3).value
                if not title_cell:
                    continue
                air_t = _parse_time(time_cell)
                if not air_t:
                    continue
                base, episode = split_title_episode(str(title_cell))
                # пропускаем строки с агрегированными эпизодами (несколько эпизодов в одном названии)
                multi_eps = len(split_base_episodes(str(title_cell))[1]) > 1
                if multi_eps and episode is None:
                    continue
                air_dt = datetime.combine(cur_date, air_t)
                duration = _parse_duration(dur_cell)
                key = (base, episode, cur_date)
                bucket = index.setdefault(key, [])
                # проверка дубликатов точного времени
                if any(existing[0] == air_dt for existing in bucket):
                    collisions += 1
                else:
                    bucket.append((air_dt, duration))
    finally:
        wb.close()
    return index, collisions

def fill_report_date_time_strict(schedule_bytes: bytes, report_path: str,
                                 sheet_name: str = 'росийские произведения',
                                 title_substr: str = 'Наименование аудиовизуального произведения',
                                 date_substr: str = 'Дата',
                                 time_substr: str = 'Время',
                                 pick_strategy: str = 'earliest') -> Dict[str, int]:
    # pick_strategy: 'earliest' | 'latest' | 'all'
    index, collisions = build_schedule_index(schedule_bytes)
    stats = {
        'matched': 0,
        'date_filled': 0,
        'time_filled': 0,
        'ambiguous': 0,
        'missed': 0,
        'collisions': collisions,
        'multi_matches': 0,
    }
    wb = load_workbook(report_path)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден")
        ws = wb[sheet_name]
        header_row = None; title_col=None; date_col=None; time_col=None
        for r in range(1, min(ws.max_row, 80)+1):
            row_vals = [str(c.value) if c.value is not None else '' for c in ws[r]]
            lowered = [v.lower() for v in row_vals]
            if any(date_substr.lower() in v for v in lowered) and any(time_substr.lower() in v for v in lowered):
                header_row = r
                for j, v in enumerate(row_vals, start=1):
                    vl = v.lower()
                    if title_substr.lower() in vl:
                        title_col = j
                    if date_substr.lower() == vl or date_substr.lower() in vl:
                        date_col = j
                    if time_substr.lower() == vl or time_substr.lower() in vl:
                        time_col = j
                break
        if not header_row or not title_col or not date_col or not time_col:
            raise RuntimeError('Не найдены шапка или необходимые колонки (Название/Дата/Время)')
        for r in range(header_row+1, ws.max_row+1):
            title_val = ws.cell(r, title_col).value
            if not title_val:
                continue
            base, episode = split_title_episode(str(title_val))
            date_val = ws.cell(r, date_col).value
            date_candidate = _parse_report_date(date_val)
            candidates: List[ValueType] = []
            if date_candidate:
                key = (base, episode, date_candidate)
                if key in index:
                    candidates = index[key]
            if not candidates:
                # без даты
                for (b, ep, d0), times_list in index.items():
                    if b == base and ep == episode:
                        candidates.extend(times_list)
            if not candidates:
                stats['missed'] += 1
                continue
            # сортируем по времени
            candidates.sort(key=lambda x: x[0])
            chosen_times: List[datetime] = []
            if len(candidates) > 1:
                stats['multi_matches'] += 1
                if pick_strategy == 'earliest':
                    chosen_times = [candidates[0][0]]
                elif pick_strategy == 'latest':
                    chosen_times = [candidates[-1][0]]
                elif pick_strategy == 'all':
                    chosen_times = [dt for dt,_ in candidates]
                else:
                    # неизвестная стратегия -> считаем ambiguous
                    stats['ambiguous'] += 1
                    continue
            else:
                chosen_times = [candidates[0][0]]
            # Заполняем. Если несколько и стратегия all -> первая дата в колонку даты, времена объединяем.
            primary_dt = chosen_times[0]
            ws.cell(r, date_col).value = primary_dt.strftime('%d.%m.%Y')
            ws.cell(r, date_col).number_format = '@'
            if pick_strategy == 'all' and len(chosen_times) > 1:
                ws.cell(r, time_col).value = '; '.join(dt.strftime('%H:%M:%S') for dt in chosen_times)
            else:
                ws.cell(r, time_col).value = primary_dt.strftime('%H:%M:%S')
            ws.cell(r, time_col).number_format = '@'
            stats['matched'] += 1
            stats['date_filled'] += 1
            stats['time_filled'] += 1
        wb.save(report_path)
    finally:
        wb.close()
    return stats
