from __future__ import annotations
from typing import Dict, Tuple, List, FrozenSet
from datetime import datetime, date, time
import re
import io

from openpyxl import load_workbook
import pandas as pd

from .normalize_titles import norm, split_base_episodes, MONTH

DATE_RE = re.compile(r'(\d{1,2})\s+([А-Яа-яЁё]+)\s+(\d{4})')
TIME_RE = re.compile(r'^(\d{1,2})[:\.](\d{2})(?::(\d{2}))?$')

MONTH_MAP = MONTH

def _parse_header_date(text: str) -> date | None:
    m = DATE_RE.search(str(text))
    if not m:
        return None
    d = int(m.group(1)); mon = MONTH_MAP.get(m.group(2).lower()); y = int(m.group(3))
    if not mon:
        return None
    return date(y, mon, d)

def _parse_time(val) -> time | None:
    if val is None:
        return None
    if isinstance(val, (pd.Timestamp, datetime)):
        return time(val.hour, val.minute, val.second)
    # excel fraction
    try:
        f = float(val)
        if -1.0 < f < 2.0:
            sec = int(round(f*24*3600))%(24*3600)
            return time(sec//3600,(sec%3600)//60, sec%60)
    except Exception:
        pass
    s = str(val).strip()
    m = TIME_RE.match(s)
    if m:
        h = int(m.group(1)); mi = int(m.group(2)); se = int(m.group(3) or 0)
        if 0<=h<24 and 0<=mi<60 and 0<=se<60:
            return time(h, mi, se)
    return None

def build_index_from_workbook(xls_bytes: bytes) -> Dict[Tuple[str, FrozenSet[int]], List[datetime]]:
    bio = io.BytesIO(xls_bytes)
    wb = load_workbook(bio, data_only=True)
    index: Dict[Tuple[str, FrozenSet[int]], List[datetime]] = {}
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            current_date: date | None = None
            for row in range(1, ws.max_row+1):
                val_time = ws.cell(row,1).value
                val_title = ws.cell(row,2).value
                # Заголовок даты во втором столбце
                d = _parse_header_date(val_title)
                if d:
                    current_date = d
                    continue
                if current_date is None:
                    continue
                t = _parse_time(val_time)
                if not t or not val_title:
                    continue
                base, episodes = split_base_episodes(str(val_title))
                key = (base, frozenset(episodes))
                index.setdefault(key, []).append(datetime.combine(current_date, t))
    finally:
        wb.close()
    return index

