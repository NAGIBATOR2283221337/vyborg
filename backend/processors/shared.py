import re, os, tempfile
from io import BytesIO
from difflib import SequenceMatcher
from typing import Dict, Tuple, Set, List, Optional
from datetime import datetime
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

# -------- ПАРАМЕТРЫ ПО УМОЛЧАНИЮ --------
DEFAULTS = dict(
    max_shows=3,
    fuzzy_cutoff=0.60,  # Снижен с 0.70 для более мягкого сопоставления
    min_token_overlap=0.40,  # Снижен с 0.50
    delete_unmatched=True  # Включено: удаляем строки без времени показа
)

TITLE_HEADER_CANDS = [
    "наименование аудиовизуального произведения",
    "наименование аудиовизуального произведения (номер и название серии)",
    "название передачи", "наименование передачи", "название программы", "наименование программы",
]
DATE_HEADER_CANDS = [
    "дата и время выхода в эфир (число, часы, мин.)",
    "дата и время выхода в эфир", "дата выхода в эфир","время выхода в эфир",
]

NOISE_TOKENS = {
    "ред","ред.","редакция","final","master","v2","v3","copy","копия","коп","сору",
    "hdrip","webrip","web","rip","bdrip","1080p","720p","uhd","4k","fullhd","hd","sd","h264","x264","x265","hevc","avc",
}

MONTHS_RU = {"января":"01","февраля":"02","марта":"03","апреля":"04","мая":"05","июня":"06",
             "июля":"07","августа":"08","сентября":"09","октября":"10","ноября":"11","декабря":"12"}

def _norm(s:str)->str:
    return re.sub(r"\s+"," ",str(s).strip().lower().replace("ё","е"))

def denoise_tokens(s: str) -> str:
    s = _norm(s)
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"\.(mp4|mov|mxf|avi|mkv)\b.*$", " ", s, flags=re.I)
    s = re.sub(r"^\s*\d{3,}[-_ ]+", " ", s)
    toks = re.split(r"[^\w]+", s)
    clean=[]
    for t in toks:
        if not t: continue
        if re.fullmatch(r"\d{3,}", t): continue
        if t in NOISE_TOKENS: continue
        clean.append(t)
    return " ".join(clean).strip(" .-–—")

def normalize_base(title:str)->str:
    base = denoise_tokens(title or "")
    base = re.sub(r"\b\d{1,3}\s*(сер(ия|ии|и)|вып(уск|уски|\.?))\b"," ",base)
    base = re.sub(r"\b(сер(ия|ии|и)|вып(уск|уски|\.?))\s*\d{1,3}\b"," ",base)
    base = re.sub(r"\b\d{1,3}\s*-\s*\d{1,3}\b"," ",base)
    return re.sub(r"\s+"," ",base).strip(" .-–—")

def extract_series_set(text:str)->Set[str]:
    s = _norm(text)
    nums=set()
    for m in re.finditer(r"\b(\d{1,3})\s*(?:-?\s*я)?\s*сер(ия|ии|и)\b",s): nums.add(m.group(1))
    for m in re.finditer(r"\bсер(ия|ии|и)\s*(\d{1,3})\b",s): nums.add(m.group(2))
    for m in re.finditer(r"\b(\d{1,3})\s*вып(уск|уски|\.?)\b",s): nums.add(m.group(1))
    for m in re.finditer(r"\bвып(уск|уски|\.?)\s*(\d{1,3})\b",s): nums.add(m.group(2))
    for m in re.finditer(r"\b(\d{1,3})\s*-\s*(\d{1,3})\s*(?:сер|вып)\b",s):
        a,b=int(m.group(1)),int(m.group(2))
        for n in range(min(a,b),max(a,b)+1): nums.add(str(n))
    for m in re.finditer(r"\b(\d{1,3})(?:\s*,\s*(\d{1,3}))+?\s*(?:сер|вып)\b",s):
        for n in re.findall(r"\d{1,3}",m.group(0)): nums.add(n)
    if not nums:
        m=re.search(r"\b(\d{1,3})\b",s)
        if m: nums.add(m.group(1))
    return nums or {"__NOSER__"}

def tokenize(s:str)->List[str]:
    return [t for t in re.split(r"[^\w]+",_norm(s)) if t]

def jaccard_over_min(a:List[str],b:List[str])->float:
    if not a or not b: return 0.0
    A=set(a); B=set(b)
    return len(A&B)/min(len(A),len(B))

def seq_ratio(a:str,b:str)->float:
    return SequenceMatcher(a=a,b=b).ratio()

def parse_time_from_str(x)->Optional[str]:
    """Извлекает время из различных форматов.

    Поддерживает:
    - Строки: "6:00", "06:00", "6:00:00"
    - Excel числа: 0.25 (= 06:00)
    - pandas Timestamp
    - datetime objects
    """
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None

    # pandas Timestamp или datetime
    if isinstance(x, (pd.Timestamp, datetime)):
        try:
            # ИСПРАВЛЕНИЕ: проверяем на NaN и преобразуем атрибуты в int
            if pd.isna(x.hour) or pd.isna(x.minute):
                return None
            hour = int(x.hour)
            minute = int(x.minute)
            return f"{hour}:{minute:02d}"
        except (ValueError, AttributeError):
            return None

    # Excel время как доля суток (0.0 - 1.0)
    if isinstance(x, (int, float)):
        try:
            f = float(x)
            # Проверяем, что это похоже на время (0.0-1.0)
            if 0.0 <= f < 1.0:
                total_seconds = int(f * 24 * 3600)
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                return f"{hours}:{minutes:02d}"
            # Если число >= 1, возможно это просто час
            elif f < 24:
                return f"{int(f)}:00"
        except:
            pass

    # Строковый формат
    try:
        s = str(x).strip()
    except:
        return None

    # Формат: "HH:MM" или "HH:MM:SS"
    m = re.match(r"^\s*(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?\s*$", s)
    if m:
        try:
            h, mi = int(m.group(1)), int(m.group(2))
            if 0 <= h < 24 and 0 <= mi < 60:
                return f"{h}:{mi:02d}"
        except:
            pass

    # Одиночное число - трактуем как час
    if re.fullmatch(r"\d{1,2}", s):
        try:
            h = int(float(s))  # На случай если "6.0"
            if 0 <= h < 24:
                return f"{h}:00"
        except:
            pass

    return None

def parse_date_label_ru(text:str)->Optional[str]:
    """Извлекает дату из текста в разных форматах.

    Поддерживает:
    - "1 сентября 2025"
    - "Понедельник, 1 сентября 2025"
    - "01.09.2025"
    - "01.09.25"
    """
    if not isinstance(text, str):
        return None

    try:
        # Формат: "1 сентября 2025" или "Понедельник, 1 сентября 2025"
        m = re.search(r"(\d{1,2})\s+([А-Яа-яЁё]+)\s+(\d{4})", text)
        if m:
            d, mon, y = m.groups()
            mon_num = MONTHS_RU.get(_norm(mon))
            if mon_num:
                # Защита: явное преобразование в int через float
                day = int(float(d))
                year = int(float(y))
                return f"{day:02d}.{mon_num}.{year}"

        # Формат: "01.09.2025"
        m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", text)
        if m:
            d, mon, y = m.groups()
            day = int(float(d))
            month = int(float(mon))
            year = int(float(y))
            return f"{day:02d}.{month:02d}.{year}"

        # Формат: "01.09.25"
        m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{2})$", text)
        if m:
            d, mon, y = m.groups()
            day = int(float(d))
            month = int(float(mon))
            year_short = int(float(y))
            full_year = f"20{year_short}" if year_short < 50 else f"19{year_short}"
            return f"{day:02d}.{month:02d}.{full_year}"
    except Exception as e:
        import logging
        logging.error(f"Ошибка парсинга даты '{text}': {e}")
        return None

    return None

def parse_dt_key(full:str):
    """Парсит строку вида '01.09.2025 в 6:00' в сортируемый ключ."""
    try:
        if not isinstance(full, str):
            return (9999, 12, 31, 23, 59)

        d, t = full.split(" в ")
        dd, mm, yyyy = d.split(".")
        h, m = t.split(":")

        # Защита: преобразуем через float на случай "01.0" и подобных
        year = int(float(yyyy))
        month = int(float(mm))
        day = int(float(dd))
        hour = int(float(h))
        minute = int(float(m))

        return (year, month, day, hour, minute)
    except Exception as e:
        import logging
        logging.debug(f"Не удалось распарсить ключ даты/времени '{full}': {e}")
        return (9999, 12, 31, 23, 59)

def limit_and_format(full_list:List[str], limit:int)->str:
    uniq=list(dict.fromkeys([x.strip() for x in full_list if x and str(x).strip()]))
    uniq.sort(key=parse_dt_key)
    return " и ".join(uniq[:limit])

def find_headers_any(ws: Worksheet, mapping=None):
    def is_title(text:str)->bool:
        t=_norm(text)
        if mapping and "title" in mapping:
            for cand in mapping["title"]:
                if _norm(cand) in t: return True
        return any(c in t for c in TITLE_HEADER_CANDS) or (("название" in t or "наименование" in t) and ("передач" in t or "произвед" in t or "программ" in t))
    def is_dt(text:str)->bool:
        t=_norm(text)
        if mapping and "aircol" in mapping:
            for cand in mapping["aircol"]:
                if _norm(cand) in t: return True
        return any(c in t for c in DATE_HEADER_CANDS) or ("дата" in t and "время" in t and "эфир" in t)

    header_row=title_col=date_col=None
    for r in range(1, min(200, ws.max_row)+1):
        for c in range(1, ws.max_column+1):
            v=ws.cell(row=r,column=c).value
            if isinstance(v,str) and is_title(v):
                header_row, title_col = r, c
                break
        if header_row: break
    if not header_row:
        raise SystemExit("Не найден столбец с названием.")

    for c in range(1, ws.max_column+1):
        v=ws.cell(row=header_row,column=c).value
        if isinstance(v,str) and is_dt(v):
            date_col=c; break
    if date_col is None:
        date_col=ws.max_column+1
        ws.cell(row=header_row,column=date_col).value="Дата и время выхода в эфир"
    return header_row, title_col, date_col

def build_schedule_index(schedule_xlsx_bytes: bytes, schedule_sheet: Optional[str]=None):
    """Читает книгу Excel из bytes, строит индекс: date -> {(base, series): [HH:MM,...]}.

    УЛУЧШЕНИЯ:
    - Автоматически ищет дату в любой колонке (не только B)
    - Поддерживает разные форматы дат
    - Ищет время и название в соседних колонках
    - Логирует процесс для отладки
    """
    import logging
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)

    mem = BytesIO(schedule_xlsx_bytes)
    xls = pd.ExcelFile(mem)

    try:
        sheet = schedule_sheet if schedule_sheet and schedule_sheet in xls.sheet_names else xls.sheet_names[0]
        logger.info(f"📖 Читаю лист: {sheet}")

        df = pd.read_excel(xls, sheet_name=sheet, header=None)
        logger.info(f"📏 Размер: {len(df)} строк × {len(df.columns)} колонок")

        rows = []
        current_date = None
        date_found_count = 0
        program_count = 0

        for idx, row in df.iterrows():
            # Ищем дату в ЛЮБОЙ колонке (не только в B)
            date_found_in_row = False
            for col_idx in range(len(row)):
                val = row.iloc[col_idx]
                if isinstance(val, str) and re.search(r"\d{4}", val):
                    parsed_date = parse_date_label_ru(val)
                    if parsed_date:
                        current_date = parsed_date
                        date_found_count += 1
                        logger.info(f"📅 Строка {int(idx)+1}: Найдена дата '{current_date}' в колонке {col_idx}")
                        date_found_in_row = True
                        break

            if date_found_in_row:
                continue

            # Теперь ищем время и название
            # Обычно: колонка 0 = время, колонка 1 = название
            # Но проверяем обе комбинации
            time_val = None
            title_val = None

            # Вариант 1: A=время, B=название
            if len(row) >= 2:
                t1 = parse_time_from_str(row.iloc[0])
                if t1 and pd.notna(row.iloc[1]):
                    time_val = t1
                    title_val = str(row.iloc[1]).strip()

            # Вариант 2: B=время, A=название (если вариант 1 не сработал)
            if not time_val and len(row) >= 2:
                t2 = parse_time_from_str(row.iloc[1])
                if t2 and pd.notna(row.iloc[0]):
                    time_val = t2
                    title_val = str(row.iloc[0]).strip()

            # Проверяем, что у нас есть всё необходимое
            if not (current_date and time_val and title_val):
                continue

            # Игнорируем служебные строки
            if len(title_val) < 3 or title_val.lower() in ['nan', 'none', '']:
                continue

            base = normalize_base(title_val)
            if not base or len(base) < 2:
                continue

            series_set = extract_series_set(title_val)
            rows.append((current_date, base, series_set, time_val))
            program_count += 1

        logger.info(f"✅ Найдено дат: {date_found_count}, программ: {program_count}")

    finally:
        xls.close()

    # Строим индекс
    schedule = {}
    for d, base, sset, t in rows:
        mp = schedule.setdefault(d, {})
        for sn in sset:
            mp.setdefault((base, sn), []).append(t)

    # Сортируем времена
    for d, mp in schedule.items():
        for k, times in mp.items():
            mp[k] = sorted(set(times), key=lambda x: (int(x.split(":")[0]), int(x.split(":")[1])))

    logger.info(f"📊 Индекс построен: {len(schedule)} дат, {sum(len(mp) for mp in schedule.values())} уникальных программ")
    return schedule
