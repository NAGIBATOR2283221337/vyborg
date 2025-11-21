# processor_rus.py – обработка российского отчёта (openpyxl-only)
from io import BytesIO
from typing import Dict
from openpyxl import load_workbook
import logging
import traceback

from .shared import (
    DEFAULTS,
    build_schedule_index,
    find_headers_any,
    limit_and_format,
)
from .matcher import pick_showtimes_for_report_title

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


def process(schedule_bytes: bytes, report_bytes: bytes, params: Dict) -> bytes:
    """Основная функция обработки одного отчёта.
    1. Строим индекс сетки (date -> (base,series)->times)
    2. Находим в отчёте строку заголовков и нужные колонки
    3. Для каждой строки отчёта ищем показы (точные + нечёткие)
    4. Заполняем колонку дат/времён, удаляем строки без совпадений если включено
    """
    try:
        p = {**DEFAULTS, **(params or {})}

        logger.info(f"🚀 Начинаю обработку с параметрами: max_shows={p['max_shows']}, "
                    f"fuzzy_cutoff={p['fuzzy_cutoff']}, min_token_overlap={p['min_token_overlap']}")

        # Индекс сетки
        logger.info("📖 Строю индекс сетки...")
        schedule = build_schedule_index(schedule_bytes, p.get("schedule_sheet"))

        # Преобразуем индекс в формат для matcher: {(base, frozenset[episodes]): [datetime, ...]}
        # schedule имеет формат: {date: {(base, series_int): [times]}}
        # Нужно преобразовать в: {(base, frozenset[episodes]): [datetime]}
        from datetime import datetime as dt

        # Сначала собираем все показы по (base, episode)
        temp_index = {}  # {(base, episode): [(date, time), ...]}

        for date_str, day_schedule in schedule.items():
            # Парсим дату из строки "DD.MM.YYYY"
            try:
                day, month, year = date_str.split('.')
                date_parts = (int(year), int(month), int(day))
            except Exception as e:
                logger.error(f"Ошибка парсинга даты '{date_str}': {e}")
                continue

            for (base, series), times in day_schedule.items():
                key = (base, series)
                if key not in temp_index:
                    temp_index[key] = []

                for time_str in times:
                    try:
                        # Парсим время из строки "HH:MM"
                        hour, minute = time_str.split(':')
                        show_datetime = dt(date_parts[0], date_parts[1], date_parts[2],
                                         int(hour), int(minute))
                        temp_index[key].append(show_datetime)
                    except Exception as e:
                        logger.error(f"Ошибка парсинга времени '{time_str}': {e}")
                        continue

        # Теперь группируем по (base, frozenset[episodes])
        matcher_index = {}

        # Собираем все эпизоды для каждой базы
        base_episodes = {}  # {base: set[episodes]}
        for (base, episode), datetimes in temp_index.items():
            if base not in base_episodes:
                base_episodes[base] = set()
            base_episodes[base].add(episode)

        # Создаем ключи для matcher
        series_count = {}  # Для статистики
        for (base, episode), datetimes in temp_index.items():
            # Для программ без серий (episode == -1) используем frozenset с -1
            if episode == -1:
                key = (base, frozenset([-1]))
            else:
                # Для программ с сериями создаем ключ только с этой серией
                key = (base, frozenset([episode]))

            if key not in matcher_index:
                matcher_index[key] = []
            matcher_index[key].extend(datetimes)

            # Статистика
            if base not in series_count:
                series_count[base] = []
            series_count[base].append(episode)

        logger.info(f"✅ Индекс построен: {len(matcher_index)} уникальных ключей")

        # Выводим детальную информацию о многосерийных программах
        multi_series = {b: eps for b, eps in series_count.items() if len(eps) > 1 and -1 not in eps}
        if multi_series:
            logger.info(f"📺 Многосерийные программы:")
            for base, episodes in sorted(multi_series.items())[:10]:  # Показываем первые 10
                episodes_sorted = sorted([e for e in episodes if e != -1])
                logger.info(f"   '{base}': серии {episodes_sorted}")

        logger.debug(f"   Примеры ключей: {list(matcher_index.keys())[:5]}")

        # Загружаем отчёт
        logger.info("📄 Загружаю отчёт...")
        wb = load_workbook(BytesIO(report_bytes))

        # Выбираем лист (по умолчанию первый, либо по имени из параметров)
        sheet_name = p.get('sheet_name')
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"📄 Используется лист: '{sheet_name}'")
        else:
            ws = wb.worksheets[0]
            logger.info(f"📄 Используется первый лист: '{ws.title}'")
        hr, tc, dc = find_headers_any(ws, p.get("mapping"))

        logger.info(f"📍 Заголовки: строка {hr}, название в колонке {tc}, даты в колонке {dc}")

        rows_to_delete = []
        matched_count = 0
        unmatched_count = 0
        total_rows = ws.max_row - hr

        for r in range(hr + 1, ws.max_row + 1):
            try:
                title_val = ws.cell(row=r, column=tc).value
                if not title_val:
                    continue

                # Показываем, что ищем
                from .normalize_titles import split_base_episodes
                search_base, search_eps = split_base_episodes(str(title_val))
                logger.info(f"🔍 Строка {r}: '{title_val}' → база='{search_base}', серии={search_eps}")

                # Используем улучшенный matcher
                found_datetimes = pick_showtimes_for_report_title(str(title_val), matcher_index)

                # Форматируем найденные времена
                if found_datetimes:
                    # Форматируем в строки "DD.MM.YYYY в HH:MM"
                    formatted_times = [
                        f"{show_dt.day:02d}.{show_dt.month:02d}.{show_dt.year} в {show_dt.hour}:{show_dt.minute:02d}"
                        for show_dt in found_datetimes
                    ]
                    formatted_value = limit_and_format(formatted_times, p["max_shows"])
                    ws.cell(row=r, column=dc).value = formatted_value
                    matched_count += 1
                    logger.info(f"✅ Строка {r}: найдено {len(found_datetimes)} показов → {formatted_value}")
                else:
                    unmatched_count += 1
                    logger.warning(f"❌ Строка {r}: '{title_val}' → не найдено совпадений")
                    if p["delete_unmatched"]:
                        rows_to_delete.append(r)
            except Exception as row_error:
                logger.error(f"❌ Ошибка обработки строки {r}: {row_error}")
                logger.error(f"   Traceback: {traceback.format_exc()}")
                # Продолжаем обработку остальных строк
                continue

        # Удаляем строки снизу вверх
        if rows_to_delete:
            logger.info(f"🗑️  Удаляю {len(rows_to_delete)} строк без совпадений...")
            for i, rr in enumerate(rows_to_delete):
                ws.delete_rows(rr - i, 1)

        logger.info(f"✅ Обработка завершена: {matched_count} совпадений, "
                    f"{unmatched_count} не найдено из {total_rows} строк")

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    except Exception as e:
        logger.error(f"💥 Критическая ошибка в process(): {e}")
        logger.error(f"   Full traceback:\n{traceback.format_exc()}")
        raise
