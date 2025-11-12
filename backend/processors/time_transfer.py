"""Модуль переноса/нормализации времени из Excel.
Задача: найти колонку времени (строка заголовков на header_row), привести значения к time,
записать форматом hh:mm:ss и сохранить новую книгу с суффиксом .filled.xlsx.
"""
from __future__ import annotations
import re
from datetime import datetime, time
from typing import Optional, Tuple, Dict
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel as excel_from_serial
import os

# Паттерны заголовков
TIME_HDR = r"^время(?:\b|\s*\()"             # Время / Время (часы, мин.)
DATE_TIME_HDR = r"дата\s*и\s*время"          # Дата и время ...
DATE_HDR = r"^дата(\b|\s*\()"               # Для доп. диагностики

__all__ = [
    "TIME_HDR", "DATE_TIME_HDR", "DATE_HDR",
    "find_col", "coerce_time", "write_time_cell", "process_time_columns"
]

# ------------------- Поиск колонки -------------------

def find_col(ws, header_row: int, pattern: str) -> Tuple[Optional[int], Optional[str]]:
    pat = re.compile(pattern, re.I)
    for c in range(1, ws.max_column + 1):
        name = str(ws.cell(header_row, c).value or "")
        if pat.search(name):
            return c, name
    return None, None

# ------------------- Приведение времени -------------------

_TIME_TOKEN = re.compile(r"^(\d{1,2})[:.](\d{1,2})(?::(\d{1,2}))?$")

def coerce_time(value) -> Optional[time]:
    """Поддержка: time, datetime, excel serial (float/int), строки HH:MM[:SS]."""
    if value is None or value == "":
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return time(value.hour, value.minute, value.second)
    if isinstance(value, (int, float)):
        try:
            dt = excel_from_serial(value)  # дата+время
            return dt.time()
        except Exception:
            pass
    s = str(value).strip()
    m = _TIME_TOKEN.match(s)
    if m:
        h = int(m.group(1)); mi = int(m.group(2)); se = int(m.group(3) or 0)
        if 0 <= h < 24 and 0 <= mi < 60 and 0 <= se < 60:
            return time(h, mi, se)
    # Попытка через strptime (избыточная, но на случай других форм)
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None

# ------------------- Запись ячейки -------------------

def write_time_cell(cell, t: time):
    # Требование: cell.value = t; cell.number_format = "hh:mm:ss"
    cell.value = t
    cell.number_format = "hh:mm:ss"

# ------------------- Основной процесс -------------------

def process_time_columns(path: str, sheet: str = "Лист1", header_row: int = 10,
                         default_time: Optional[time] = None) -> Dict[str, int | str]:
    """Нормализует колонку времени. Возвращает телеметрию.

    Шаги:
      1) Найти колонку времени по TIME_HDR; если нет – объединённую по DATE_TIME_HDR.
      2) Подсчитать пустые (coerce_time == None) до.
      3) Переписать каждую ячейку в формат hh:mm:ss (если есть значение или default_time).
      4) Подсчитать пустые после.
      5) Сохранить книгу под новым именем <path>.filled.xlsx.
    """
    wb = load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Лист '{sheet}' не найден")
    ws = wb[sheet]
    c_time, hdr_time = find_col(ws, header_row, TIME_HDR)
    c_dt, hdr_dt = find_col(ws, header_row, DATE_TIME_HDR)
    source_col = c_time or c_dt
    if not source_col:
        wb.close()
        raise RuntimeError("Не найдена колонка времени по шаблонам TIME_HDR / DATE_TIME_HDR")

    before_empty = 0
    for r in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(r, source_col).value
        if coerce_time(raw) is None:
            before_empty += 1

    filled = 0
    after_empty = 0
    for r in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(r, source_col)
        raw = cell.value
        t = coerce_time(raw)
        if t is None:
            if isinstance(default_time, time):
                write_time_cell(cell, default_time)
                filled += 1
            else:
                continue
        else:
            write_time_cell(cell, t)
    # после заполнения считаем пустые снова
    for r in range(header_row + 1, ws.max_row + 1):
        if coerce_time(ws.cell(r, source_col).value) is None:
            after_empty += 1

    out_path = path.replace('.xlsx', '.filled.xlsx')
    wb.save(out_path)
    wb.close()
    return {
        'before_time_empty': before_empty,
        'after_time_empty': after_empty,
        'filled_cells': filled,
        'saved_to': out_path,
        'time_column_index': source_col,
        'time_header': hdr_time or hdr_dt,
    }

