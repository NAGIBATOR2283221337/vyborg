from __future__ import annotations
from typing import Dict, Tuple, List, FrozenSet
from datetime import datetime
import logging

from openpyxl import load_workbook

from .schedule_index import build_index_from_workbook
from .matcher import pick_showtimes_for_report_title, best_candidates
from .normalize_titles import split_base_episodes


logger = logging.getLogger(__name__)
if not logger.handlers:
    h = logging.StreamHandler()
    fmt = logging.Formatter('[MATCH] %(message)s')
    h.setFormatter(fmt)
    logger.addHandler(h)
    logger.setLevel(logging.INFO)


def match_times_for_title(title: str, index: Dict[Tuple[str,FrozenSet[int]], List[datetime]]) -> List[datetime]:
    base, episodes = split_base_episodes(title)
    key_full = (base, frozenset(episodes))
    if key_full in index:
        return index[key_full]
    out: List[datetime] = []
    for (b, eps_set), times in index.items():
        if b != base:
            continue
        if not episodes and not eps_set:
            out.extend(times)
        elif episodes and (eps_set & episodes):
            out.extend(times)
    return sorted(set(out))


def format_dt_list(dts: List[datetime]) -> str:
    uniq = sorted(set(dts))
    return "; ".join(dt.strftime("%d.%m.%Y %H:%M") for dt in uniq)


def _find_header_row(ws, target_substr: str) -> int | None:
    target_low = target_substr.lower()
    for r in range(1, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            val = ws.cell(r,c).value
            if val and target_low in str(val).lower():
                return r
    return None


def _find_col_by_substr(ws, header_row: int, substr: str) -> int | None:
    substr_low = substr.lower()
    for c in range(1, ws.max_column+1):
        val = ws.cell(header_row, c).value
        if val and substr_low in str(val).lower():
            return c
    return None


def fill_report_column(schedule_xls_bytes: bytes, report_path: str,
                        sheet_name: str = "росийские произведения",
                        title_substr: str = "Наименование аудиовизуального произведения",
                        target_substr: str = "Дата и время выхода в эфир (число, часы, мин.)") -> None:
    index = build_index_from_workbook(schedule_xls_bytes)
    wb = load_workbook(report_path)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден в отчёте")
        ws = wb[sheet_name]
        header_row = _find_header_row(ws, target_substr)
        if header_row is None:
            raise ValueError("Строка шапки не найдена")
        title_col = _find_col_by_substr(ws, header_row, title_substr)
        target_col = _find_col_by_substr(ws, header_row, target_substr)
        if not title_col or not target_col:
            raise ValueError("Не найдены нужные колонки")
        for r in range(header_row+1, ws.max_row+1):
            cell_title = ws.cell(r, title_col).value
            if not cell_title:
                continue
            dts = pick_showtimes_for_report_title(str(cell_title), index)
            if not dts:
                cands, _ = best_candidates(str(cell_title), index.keys())
                logger.info(f"NO MATCH: '{cell_title}' -> candidates: {cands[:3]}")
                continue
            cell = ws.cell(r, target_col)
            cell.value = format_dt_list(dts)
            cell.number_format = '@'
        wb.save(report_path)
    finally:
        wb.close()


def fill_report_column_and_prune(schedule_xls_bytes: bytes, report_path: str,
                        sheet_name: str = "росийские произведения",
                        title_substr: str = "Наименование аудиовизуального произведения",
                        target_substr: str = "Дата и время выхода в эфир (число, часы, мин.)") -> None:
    index = build_index_from_workbook(schedule_xls_bytes)
    wb = load_workbook(report_path)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден в отчёте")
        ws = wb[sheet_name]
        # Поиск шапки
        header_row = None; title_col=None; target_col=None
        for r in range(1, min(ws.max_row, 60)+1):
            row_vals=[str(c.value) if c.value is not None else "" for c in ws[r]]
            if any("Дата и время выхода" in v for v in row_vals):
                header_row = r
                for j,v in enumerate(row_vals, start=1):
                    if title_substr in v: title_col=j
                    if "Дата и время выхода" in v: target_col=j
                break
        if not header_row or not title_col or not target_col:
            raise RuntimeError("Не удалось найти шапку и нужные колонки")
        rows_to_delete = []
        for r in range(header_row+1, ws.max_row+1):
            title = ws.cell(r, title_col).value
            if not title:
                continue
            dts = pick_showtimes_for_report_title(str(title), index)
            if not dts:
                rows_to_delete.append(r)
                continue
            cell = ws.cell(r, target_col)
            cell.value = "; ".join(sorted({dt.strftime("%d.%m.%Y %H:%M") for dt in dts}))
            cell.number_format = "@"
        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r, 1)
        wb.save(report_path)
    finally:
        wb.close()
