import os
import tempfile
import gc
from io import BytesIO
from typing import Dict

import openpyxl
from openpyxl import load_workbook

try:
    from .shared import (
        ensure_real_xlsx,
        build_schedule_index,
        find_headers_any,
        parse_date_from_cell,
        normalize_base,
        tokenize,
        jaccard_over_min,
        seq_ratio,
        limit_and_format
    )
except ImportError:
    from shared import (
        ensure_real_xlsx,
        build_schedule_index,
        find_headers_any,
        parse_date_from_cell,
        normalize_base,
        tokenize,
        jaccard_over_min,
        seq_ratio,
        limit_and_format
    )


def process(schedule_bytes: bytes, report_bytes: bytes, params: Dict) -> bytes:
    """
    Обработка российского отчёта

    Args:
        schedule_bytes: Байты файла сетки
        report_bytes: Байты файла отчёта
        params: Параметры обработки

    Returns:
        bytes: Готовый xlsx файл
    """
    max_shows = params.get('max_shows', 3)
    fuzzy_cutoff = params.get('fuzzy_cutoff', 0.20)
    min_token_overlap = params.get('min_token_overlap', 0.35)
    delete_unmatched = params.get('delete_unmatched', True)

    # Создаем временную директорию
    with tempfile.TemporaryDirectory() as temp_dir:
        # Сохраняем входные файлы
        schedule_path = os.path.join(temp_dir, "schedule.xlsx")
        report_path = os.path.join(temp_dir, "report.xlsx")

        with open(schedule_path, 'wb') as f:
            f.write(schedule_bytes)

        with open(report_path, 'wb') as f:
            f.write(report_bytes)

        # Конвертируем в xlsx если нужно
        schedule_path = ensure_real_xlsx(schedule_path)
        report_path = ensure_real_xlsx(report_path)

        # Индексируем сетку
        schedule_by_date, bases_by_date = build_schedule_index(schedule_path)

        # Обрабатываем отчёт
        wb = None
        try:
            wb = load_workbook(report_path)
            ws = wb.active

            # Находим заголовки
            header_row, title_col, date_col = find_headers_any(ws)

            # Список строк для удаления (индексы)
            rows_to_delete = []

            # Обрабатываем строки отчёта
            for r in range(header_row + 1, ws.max_row + 1):
                title_cell = ws.cell(r, title_col)
                date_cell = ws.cell(r, date_col)

                title_val = title_cell.value
                cell_val = date_cell.value

                if not title_val:
                    continue

                # Парсим дату
                date_r = parse_date_from_cell(cell_val)

                if not date_r:
                    if delete_unmatched:
                        rows_to_delete.append(r)
                    continue

                # Получаем данные сетки для этой даты
                sub = schedule_by_date.get(date_r)
                if sub is None or sub.empty:
                    if delete_unmatched:
                        rows_to_delete.append(r)
                    continue

                # Нормализуем название из отчёта
                base_r = normalize_base(str(title_val))
                if not base_r:
                    if delete_unmatched:
                        rows_to_delete.append(r)
                    continue

                # Токенизируем
                tokens_r = tokenize(base_r)

                # Ищем лучшее совпадение
                best_b = None
                best_score = 0.0

                for _, row_data in sub.iterrows():
                    base_s = row_data['base']
                    tokens_s = tokenize(base_s)

                    # Вычисляем метрики
                    overlap = jaccard_over_min(tokens_r, tokens_s)
                    ratio = seq_ratio(base_r, base_s)

                    # Проверяем пороги
                    if ratio >= fuzzy_cutoff or overlap >= min_token_overlap:
                        score = max(overlap, ratio)
                        if score > best_score:
                            best_score = score
                            best_b = base_s

                if best_b is None:
                    if delete_unmatched:
                        rows_to_delete.append(r)
                    continue

                # Собираем времена для лучшего совпадения
                matching_rows = sub[sub['base'] == best_b]
                times = matching_rows['time'].tolist()

                # Форматируем показы
                formatted_shows = []
                for time_val in times:
                    formatted_shows.append(f"{date_r} в {time_val}")

                # Применяем лимит и форматирование
                result_text = limit_and_format(formatted_shows, max_shows)

                # Записываем результат в ячейку даты
                date_cell.value = result_text

            # Удаляем строки (с конца, чтобы не сбить индексы)
            for row_idx in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_idx)

            # Сохраняем в память
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        finally:
            # Обязательно закрываем workbook
            if wb is not None:
                wb.close()
            # Принудительно освобождаем память
            gc.collect()

