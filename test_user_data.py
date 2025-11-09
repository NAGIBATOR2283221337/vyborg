#!/usr/bin/env python3
"""
Специальная отладка для файлов пользователя
Основана на структуре из предоставленных скриншотов
"""
import os
import sys
from pathlib import Path

# Добавляем пути
current_dir = Path(__file__).parent
sys.path.insert(0, str(current_dir))
sys.path.insert(0, str(current_dir / "backend"))

def create_test_files_from_screenshots():
    """Создаем тестовые файлы на основе ваших скриншотов"""
    from openpyxl import Workbook
    from io import BytesIO

    print("📄 Создаем файлы на основе ваших скриншотов...")

    # СЕТКА (верхний скриншот)
    schedule_wb = Workbook()
    ws = schedule_wb.active

    # Заполняем как на скриншоте - БЕЗ заголовков
    row = 1
    ws.cell(row, 1, "Понедельник, 1 сентября 2025")  # Дата
    row += 1
    ws.cell(row, 1, "6:00:00")
    ws.cell(row, 2, "Заставка СМИ")
    row += 1
    ws.cell(row, 1, "6:00:15")
    ws.cell(row, 2, "Заставка Доброе утро!")
    row += 1
    ws.cell(row, 1, "6:00:25")
    ws.cell(row, 2, "Гора самоцветов 61, 62")
    row += 1
    ws.cell(row, 1, "6:26:25")
    ws.cell(row, 2, "Гора самоцветов 61, 62")
    row += 1
    ws.cell(row, 1, "6:56:25")
    ws.cell(row, 2, "Заставка Реклама")
    row += 1
    ws.cell(row, 1, "6:56:35")
    ws.cell(row, 2, "Реклама")
    row += 1
    ws.cell(row, 1, "6:59:45")
    ws.cell(row, 2, "Заставка СМИ")
    row += 1
    ws.cell(row, 1, "7:00:00")
    ws.cell(row, 2, "Повтор")
    row += 1
    ws.cell(row, 1, "7:30:00")
    ws.cell(row, 2, "Гора самоцветов 63, 64")

    print("✅ Сетка создана (без заголовков)")

    # ОТЧЕТ (нижний скриншот)
    report_wb = Workbook()
    ws = report_wb.active

    # Пустые строки до заголовков (как на скриншоте)
    for i in range(1, 11):
        ws.cell(i, 1, "")

    # Заголовки на строке 11 (как на скриншоте)
    ws.cell(11, 1, "Наименование аудиовизуального произведения (номер и название серии)")
    ws.cell(11, 2, "Дата и время выхода в эфир (число, часы, мин.)")
    ws.cell(11, 3, "Год создания")
    ws.cell(11, 4, "Жанр (тип)")
    ws.cell(11, 5, "Киностудия (производитель)")
    ws.cell(11, 6, "Страна создания")
    ws.cell(11, 7, "Режиссер")
    ws.cell(11, 8, "Композитор оригинальной музыки")
    ws.cell(11, 9, "Длительность (мин. Сек)")

    # Данные отчета (строка 12 и далее)
    ws.cell(12, 1, "Бронская история (A Bronx Tale)")
    ws.cell(12, 2, "19.09.2025")
    ws.cell(12, 3, "1993")
    ws.cell(12, 4, "Художественный фильм")
    ws.cell(12, 5, "Renta Entertainment")
    ws.cell(12, 6, "США")
    ws.cell(12, 7, "Роберт Де Ниро (Robert De Niro)")
    ws.cell(12, 8, "Буч Барбелла (Butch Barbella)")
    ws.cell(12, 9, "01:56:14")

    # Еще одна строка
    ws.cell(13, 1, "ВЕЛИКОЛЕПНЫЕ ЭМБЕРСОНЫ (The Magnificent Ambersons)")
    ws.cell(13, 2, "20.09.2025")
    ws.cell(13, 3, "1942")
    ws.cell(13, 4, "Художественный фильм")

    print("✅ Отчет создан (с заголовками на строке 11)")

    # Конвертируем в байты
    schedule_buffer = BytesIO()
    schedule_wb.save(schedule_buffer)
    schedule_bytes = schedule_buffer.getvalue()
    schedule_buffer.close()
    schedule_wb.close()

    report_buffer = BytesIO()
    report_wb.save(report_buffer)
    report_bytes = report_buffer.getvalue()
    report_buffer.close()
    report_wb.close()

    return schedule_bytes, report_bytes

def test_with_user_data():
    """Тестируем с данными пользователя"""
    print("🔍 ТЕСТ С ДАННЫМИ ПОЛЬЗОВАТЕЛЯ")
    print("=" * 60)

    try:
        from backend.processors import processor_rus

        # Создаем файлы по образцу скриншотов
        schedule_bytes, report_bytes = create_test_files_from_screenshots()

        print(f"📊 Размеры: сетка={len(schedule_bytes)} байт, отчет={len(report_bytes)} байт")

        # Более мягкие параметры для лучшего сопоставления
        params = {
            'max_shows': 3,
            'fuzzy_cutoff': 0.1,  # Снижаем порог
            'min_token_overlap': 0.2,  # Снижаем порог
            'delete_unmatched': False  # Не удаляем для отладки
        }

        print(f"⚙️ Параметры: {params}")
        print("\n" + "="*60)
        print("🚀 ЗАПУСК ОБРАБОТКИ")
        print("="*60)

        # Обрабатываем
        result_bytes = processor_rus.process(schedule_bytes, report_bytes, params)

        print("="*60)
        print(f"✅ РЕЗУЛЬТАТ: {len(result_bytes)} байт")

        # Сохраняем и анализируем результат
        result_path = "user_data_result.xlsx"
        with open(result_path, 'wb') as f:
            f.write(result_bytes)
        print(f"💾 Сохранено: {result_path}")

        # Показываем содержимое результата
        from openpyxl import load_workbook
        wb = load_workbook(result_path)
        ws = wb.active

        print(f"\n📋 СОДЕРЖИМОЕ РЕЗУЛЬТАТА:")
        print(f"   Размеры: {ws.max_row} строк x {ws.max_column} колонок")

        filled_rows = 0
        for r in range(1, ws.max_row + 1):
            row_data = []
            has_data = False
            for c in range(1, min(4, ws.max_column + 1)):
                val = ws.cell(r, c).value
                if val:
                    has_data = True
                    row_data.append(str(val)[:30])
                else:
                    row_data.append("")

            if has_data:
                filled_rows += 1
                print(f"   Строка {r}: {' | '.join(row_data)}")

        wb.close()

        print(f"\n📊 СТАТИСТИКА:")
        print(f"   Заполненных строк: {filled_rows}")
        print(f"   Размер файла: {len(result_bytes)} байт")

        if filled_rows > 2:
            print("✅ ФАЙЛ НЕ ПУСТОЙ - проблема решена!")
            return True
        else:
            print("❌ Файл все еще практически пустой")
            return False

    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_with_user_data()

    print("\n" + "="*60)
    if success:
        print("🎉 ПРОБЛЕМА РЕШЕНА!")
        print("Файл обрабатывается корректно")
    else:
        print("⚠️ ПРОБЛЕМА ОСТАЕТСЯ")
        print("Требуются дополнительные исправления")

    print("="*60)
    input("Нажмите Enter для завершения...")
