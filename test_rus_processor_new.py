import io
from datetime import datetime
import pandas as pd
from backend.processors.processor_rus import process

# Помощник для создания xlsx в памяти
from openpyxl import Workbook

def make_schedule():
    wb = Workbook(); ws = wb.active
    ws.cell(1,1).value = 'Понедельник, 1 сентября 2025'
    ws.cell(2,1).value = '06:00'; ws.cell(2,2).value = 'Новости'
    ws.cell(3,1).value = '07:00'; ws.cell(3,2).value = 'Новости'
    ws.cell(4,1).value = '08:00'; ws.cell(4,2).value = 'Гора самоцветов 63,64'
    ws.cell(5,1).value = '09:00'; ws.cell(5,2).value = 'Гора самоцветов 63,64'
    bio = io.BytesIO(); wb.save(bio); bio.seek(0); return bio.getvalue()

def make_report():
    wb = Workbook(); ws = wb.active
    # заполняем 11 пустых строк для header=11
    for r in range(1,12):
        ws.cell(r,1).value = ''
    ws.cell(12,1).value = 'Наименование аудиовизуального произведения (номер и название серии)'
    ws.cell(12,2).value = 'Дата и время выхода в эфир (число, часы, мин.)'
    ws.cell(13,1).value = 'Новости'
    ws.cell(14,1).value = 'Гора самоцветов. 63 серия'
    ws.cell(15,1).value = 'Несуществующая передача'
    bio = io.BytesIO(); wb.save(bio); bio.seek(0); return bio.getvalue()

def test_process_basic():
    schedule_bytes = make_schedule()
    report_bytes = make_report()
    out = process(schedule_bytes, report_bytes, {'max_shows':50})
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(out)); ws = wb.active
    # Заголовок смещён: 12 строк шапка
    assert ws.cell(13,2).value.startswith('01.09.2025')  # Новости получили даты-время
    assert '06:00' in ws.cell(13,2).value
    assert '07:00' in ws.cell(13,2).value
    # Гора самоцветов 63 серия должна взять первое время 08:00
    assert '08:00' in ws.cell(14,2).value
    # Несуществующая пусто
    assert ws.cell(15,2).value == '' or ws.cell(15,2).value is None

if __name__ == '__main__':
    test_process_basic()
    print('✅ Новый процессор прошёл базовый тест')

