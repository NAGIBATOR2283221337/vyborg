"""Диагностика структуры Excel файлов для отладки обработки."""
import pandas as pd
import openpyxl
from pathlib import Path

def diagnose_excel(filepath: str, max_rows: int = 30):
    """Показывает структуру Excel файла."""
    print(f"\n{'='*80}")
    print(f"ФАЙЛ: {filepath}")
    print(f"{'='*80}\n")

    # Список листов
    wb = openpyxl.load_workbook(filepath)
    print(f"📄 Листы в файле: {wb.sheetnames}\n")

    for sheet_name in wb.sheetnames[:3]:  # первые 3 листа
        print(f"\n{'─'*80}")
        print(f"📊 ЛИСТ: {sheet_name}")
        print(f"{'─'*80}\n")

        ws = wb[sheet_name]
        print(f"Размер: {ws.max_row} строк × {ws.max_column} колонок\n")

        # Показываем первые строки
        print(f"Первые {min(max_rows, ws.max_row)} строк:\n")
        print(f"{'Строка':<8} | {'A':<30} | {'B':<30} | {'C':<30}")
        print(f"{'-'*8}-+-{'-'*30}-+-{'-'*30}-+-{'-'*30}")

        for r in range(1, min(max_rows + 1, ws.max_row + 1)):
            a_val = ws.cell(r, 1).value
            b_val = ws.cell(r, 2).value if ws.max_column >= 2 else None
            c_val = ws.cell(r, 3).value if ws.max_column >= 3 else None

            # Форматируем значения
            a_str = str(a_val)[:28] if a_val else ""
            b_str = str(b_val)[:28] if b_val else ""
            c_str = str(c_val)[:28] if c_val else ""

            # Показываем тип данных
            a_type = f"({type(a_val).__name__})" if a_val else ""
            b_type = f"({type(b_val).__name__})" if b_val else ""
            c_type = f"({type(c_val).__name__})" if c_val else ""

            print(f"{r:<8} | {a_str:<30} | {b_str:<30} | {c_str:<30}")
            if any([a_type, b_type, c_type]):
                print(f"{'':8} | {a_type:<30} | {b_type:<30} | {c_type:<30}")

        print()

    wb.close()


if __name__ == "__main__":
    # Путь к тестовым файлам
    test_dir = Path(__file__).parent / "tests"

    # Диагностируем сетку
    schedule_file = test_dir / "Копия Сентябрь в работе.xlsx"
    if schedule_file.exists():
        diagnose_excel(str(schedule_file), max_rows=50)
    else:
        print(f"❌ Файл не найден: {schedule_file}")

    # Если есть отчётные файлы, тоже диагностируем
    report_files = list(test_dir.glob("*отчет*.xlsx")) + list(test_dir.glob("*report*.xlsx"))
    for report_file in report_files[:2]:  # первые 2
        diagnose_excel(str(report_file), max_rows=20)

