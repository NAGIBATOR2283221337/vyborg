"""Быстрая проверка структуры файла сетки."""
import openpyxl

filepath = r"C:\Users\User\Desktop\PythonProject\PythonProject\tests\Копия Сентябрь в работе.xlsx"

try:
    wb = openpyxl.load_workbook(filepath)
    print(f"Листы: {wb.sheetnames}\n")

    ws = wb[wb.sheetnames[0]]
    print(f"Первый лист: {wb.sheetnames[0]}")
    print(f"Размер: {ws.max_row} строк x {ws.max_column} колонок\n")
    print("Первые 30 строк (A, B, C):\n")

    for r in range(1, min(31, ws.max_row + 1)):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        print(f"{r:3}. A={str(a)[:25]:<25} | B={str(b)[:35]:<35} | C={str(c)[:20] if c else ''}")

    wb.close()
except Exception as e:
    print(f"Ошибка: {e}")
    import traceback
    traceback.print_exc()

